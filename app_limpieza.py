import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════
#  LÓGICA DE LIMPIEZA
# ══════════════════════════════════════════════════════════

def normalizar_nombre(nombre):
    return re.sub(r"\s+", "_", str(nombre).strip().lower())

def limpiar_archivo(ruta_entrada, log):
    nombre_base = os.path.splitext(os.path.basename(ruta_entrada))[0]
    ruta_salida = os.path.join(os.path.dirname(ruta_entrada), f"{nombre_base}_limpio.xlsx")

    log("Leyendo archivo...")
    df = pd.read_excel(ruta_entrada, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    filas_originales = len(df)
    log(f"✔ {filas_originales} registros encontrados")
    log(f"✔ Columnas: {', '.join(df.columns)}")

    # Filas vacías
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    vacias = filas_originales - len(df)
    if vacias:
        log(f"✔ {vacias} filas vacías eliminadas")

    palabras_fecha  = ["fecha", "date", "ingreso", "nacimiento", "registro"]
    palabras_numero = ["salario", "sueldo", "monto", "precio", "total", "cantidad", "id"]
    palabras_bool   = ["activo", "estado", "vigente", "habilitado"]
    palabras_email  = ["email", "correo", "mail", "e-mail"]

    cols_email = [c for c in df.columns if any(p in normalizar_nombre(c) for p in palabras_email)]
    for col in cols_email:
        df[col] = df[col].apply(
            lambda v: str(v).strip().lower() if pd.notna(v) and str(v).strip() != "" else v)
    if cols_email:
        log(f"✔ Correos normalizados en: {', '.join(cols_email)}")

    def limpiar_texto(v):
        if pd.isna(v) or str(v).strip() == "": return v
        return " ".join(str(v).split()).title()

    cols_excluir = palabras_fecha + palabras_numero + palabras_email
    cols_texto = [c for c in df.columns
                  if not any(p in normalizar_nombre(c) for p in cols_excluir)]
    for col in cols_texto:
        df[col] = df[col].apply(limpiar_texto)
    log(f"✔ Texto normalizado en: {', '.join(cols_texto)}")

    cols_fecha = [c for c in df.columns
                  if any(p in normalizar_nombre(c) for p in palabras_fecha)]
    for col in cols_fecha:
        cnt = [0]
        def pf(v, c=cnt):
            if pd.isna(v) or str(v).strip() == "": return None
            try:
                f = pd.to_datetime(v, dayfirst=True)
                c[0] += 1
                return f.to_pydatetime()
            except: return v
        df[col] = df[col].apply(pf)
        log(f"✔ Fechas reales en '{col}': {cnt[0]} valores")

    cols_num = [c for c in df.columns
                if any(p in normalizar_nombre(c) for p in palabras_numero)]
    for col in cols_num:
        def ln(v):
            if pd.isna(v): return None
            limpio = re.sub(r"[$\.\s]", "", str(v)).replace(",", ".")
            try: return int(float(limpio))
            except: return v
        df[col] = df[col].apply(ln)
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if cols_num:
        log(f"✔ Números limpiados en: {', '.join(cols_num)}")

    mapa_bool = {"si":"Sí","sí":"Sí","yes":"Sí","true":"Sí","1":"Sí","verdadero":"Sí",
                 "no":"No","not":"No","false":"No","0":"No","falso":"No"}
    cols_bool = [c for c in df.columns
                 if any(p in normalizar_nombre(c) for p in palabras_bool)]
    for col in cols_bool:
        df[col] = df[col].apply(
            lambda v: mapa_bool.get(str(v).strip().lower(), v) if pd.notna(v) else v)
        log(f"✔ '{col}' normalizado")

    antes = len(df)
    df.drop_duplicates(inplace=True)
    df.reset_index(drop=True, inplace=True)
    dup = antes - len(df)
    log(f"✔ {dup} duplicados eliminados" if dup else "✔ Sin duplicados")

    # Exportar con formato
    log("Exportando archivo...")
    df.to_excel(ruta_salida, index=False, sheet_name="Datos Limpios")
    wb = load_workbook(ruta_salida)
    ws = wb.active
    C_ENC, C_PAR, C_BRD = "2E4057", "F0F4F8", "B0BEC5"
    borde = Border(**{l: Side(style="thin", color=C_BRD)
                      for l in ["left","right","top","bottom"]})

    for cell in ws[1]:
        cell.font      = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color=C_ENC)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = borde

    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        bg = C_PAR if ri % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font      = Font(name="Segoe UI", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = borde
            # Formato de fecha en columnas detectadas
            if cell.column_letter in [
                get_column_letter(ws[1][i].column)
                for i, h in enumerate(ws[1])
                if any(p in normalizar_nombre(str(h.value or "")) for p in palabras_fecha)
            ]:
                cell.number_format = "YYYY-MM-DD"

    for col in ws.columns:
        mx = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx + 4
    ws.row_dimensions[1].height = 22

    ws2 = wb.create_sheet("Resumen")
    datos_res = [
        ("Resumen", ""),
        ("Registros originales",       filas_originales),
        ("Filas vacías eliminadas",     vacias),
        ("Duplicados eliminados",       dup),
        ("Registros finales",           len(df)),
        ("Columnas procesadas",         len(df.columns)),
    ]
    for i, (k, v) in enumerate(datos_res, 1):
        ws2.cell(i, 1, k)
        ws2.cell(i, 2, str(v))
        for j in [1, 2]:
            c = ws2.cell(i, j)
            c.font      = Font(name="Segoe UI", bold=(i==1), size=10,
                               color="FFFFFF" if i==1 else "000000")
            c.fill      = PatternFill("solid", start_color=C_ENC if i==1
                                      else (C_PAR if i%2==0 else "FFFFFF"))
            c.border    = borde
            c.alignment = Alignment(horizontal="left")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    wb.save(ruta_salida)

    return ruta_salida, filas_originales, len(df), dup, vacias

# ══════════════════════════════════════════════════════════
#  INTERFAZ GRÁFICA
# ══════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Limpieza de Excel")
        self.geometry("560x620")
        self.resizable(False, False)
        self.configure(bg="#0F1923")
        self.ruta_archivo = None
        self._build_ui()

    def _build_ui(self):
        header = tk.Frame(self, bg="#0F1923")
        header.pack(fill="x", padx=30, pady=(30, 0))

        tk.Label(header, text="✦", font=("Segoe UI", 22), bg="#0F1923", fg="#4ECDC4").pack(anchor="w")
        tk.Label(header, text="Limpieza\nde Datos", font=("Georgia", 26, "bold"),
                 bg="#0F1923", fg="#F5F0E8", justify="left").pack(anchor="w")
        tk.Label(header, text="Sube tu Excel y lo dejamos impecable.",
                 font=("Segoe UI", 10), bg="#0F1923", fg="#6B7F8E").pack(anchor="w", pady=(4,0))

        tk.Frame(self, bg="#1E2D3D", height=1).pack(fill="x", padx=30, pady=20)

        self.frame_archivo = tk.Frame(self, bg="#141F2B", bd=0, relief="flat",
                                      highlightbackground="#2A3F52", highlightthickness=1)
        self.frame_archivo.pack(fill="x", padx=30, pady=(0, 16))

        inner = tk.Frame(self.frame_archivo, bg="#141F2B")
        inner.pack(padx=20, pady=16, fill="x")

        tk.Label(inner, text="📄  Archivo Excel", font=("Segoe UI", 9, "bold"),
                 bg="#141F2B", fg="#4ECDC4").pack(anchor="w")

        self.lbl_archivo = tk.Label(inner, text="Ningún archivo seleccionado",
                                    font=("Segoe UI", 10), bg="#141F2B", fg="#6B7F8E",
                                    anchor="w", wraplength=440)
        self.lbl_archivo.pack(anchor="w", pady=(4, 12), fill="x")

        tk.Button(inner, text="  Seleccionar archivo",
                  font=("Segoe UI", 10, "bold"), bg="#1E2D3D", fg="#F5F0E8",
                  activebackground="#2A3F52", activeforeground="#4ECDC4",
                  bd=0, padx=16, pady=8, cursor="hand2",
                  command=self.seleccionar_archivo).pack(anchor="w")

        self.btn_procesar = tk.Button(self, text="⚡  Procesar",
                                      font=("Segoe UI", 12, "bold"),
                                      bg="#4ECDC4", fg="#0F1923",
                                      activebackground="#3DBDB5", activeforeground="#0F1923",
                                      bd=0, padx=0, pady=14, cursor="hand2",
                                      state="disabled", command=self.procesar)
        self.btn_procesar.pack(fill="x", padx=30, pady=(0, 16))

        frame_log = tk.Frame(self, bg="#0F1923")
        frame_log.pack(fill="both", expand=True, padx=30, pady=(0, 20))

        tk.Label(frame_log, text="Registro de operaciones",
                 font=("Segoe UI", 9, "bold"), bg="#0F1923", fg="#4ECDC4").pack(anchor="w", pady=(0,6))

        self.log_box = tk.Text(frame_log, height=12, font=("Consolas", 9),
                               bg="#0A1219", fg="#A8C0CC", insertbackground="#4ECDC4",
                               bd=0, relief="flat", padx=12, pady=10,
                               highlightbackground="#1E2D3D", highlightthickness=1,
                               state="disabled")
        self.log_box.pack(fill="both", expand=True)

        self.lbl_estado = tk.Label(self, text="Esperando archivo...",
                                   font=("Segoe UI", 9), bg="#0F1923", fg="#3D5166")
        self.lbl_estado.pack(pady=(0, 16))

    def seleccionar_archivo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if ruta:
            self.ruta_archivo = ruta
            nombre = os.path.basename(ruta)
            self.lbl_archivo.config(text=f"📂  {nombre}", fg="#F5F0E8")
            self.btn_procesar.config(state="normal")
            self.set_estado("Archivo listo. Haz clic en Procesar.")
            self.escribir_log(f"Archivo seleccionado: {nombre}\n")

    def procesar(self):
        if not self.ruta_archivo:
            return
        self.btn_procesar.config(state="disabled", text="Procesando...")
        self.set_estado("Procesando...")
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

        def tarea():
            try:
                ruta_out, orig, final, dup, vacias = limpiar_archivo(
                    self.ruta_archivo, self.escribir_log)
                self.after(0, lambda: self.mostrar_exito(ruta_out, orig, final, dup, vacias))
            except Exception as e:
                self.after(0, lambda: self.mostrar_error(str(e)))

        threading.Thread(target=tarea, daemon=True).start()

    def mostrar_exito(self, ruta, orig, final, dup, vacias):
        self.btn_procesar.config(state="normal", text="⚡  Procesar")
        self.set_estado(f"✔ Listo  —  {orig} → {final} registros")
        self.escribir_log(f"\n{'─'*40}")
        self.escribir_log(f"✔ PROCESO COMPLETADO")
        self.escribir_log(f"  Registros originales : {orig}")
        self.escribir_log(f"  Filas vacías          : {vacias}")
        self.escribir_log(f"  Duplicados eliminados : {dup}")
        self.escribir_log(f"  Registros finales     : {final}")
        self.escribir_log(f"  Archivo guardado en:")
        self.escribir_log(f"  {ruta}")
        messagebox.showinfo("¡Listo!", f"Archivo generado correctamente:\n\n{os.path.basename(ruta)}")

    def mostrar_error(self, msg):
        self.btn_procesar.config(state="normal", text="⚡  Procesar")
        self.set_estado("✖ Error al procesar")
        self.escribir_log(f"\n✖ ERROR: {msg}")
        messagebox.showerror("Error", f"Ocurrió un problema:\n\n{msg}")

    def escribir_log(self, texto):
        self.log_box.config(state="normal")
        self.log_box.insert("end", texto + "\n")
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    def set_estado(self, texto):
        self.lbl_estado.config(text=texto)


if __name__ == "__main__":
    app = App()
    app.mainloop()
